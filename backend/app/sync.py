"""Sincronizador: lê as planilhas FLUXO DE CAIXA, aplica os filtros de negócio
(RN-01, RN-08, RN-04) e grava tudo no SQLite via db.py.

As planilhas nunca são escritas (RN-07) — abertas apenas em modo leitura.
"""
from __future__ import annotations

import logging
import re
import time
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from . import db
from .config import Config, carregar_config

logger = logging.getLogger("consulta_recibos.sync")

COLUNAS_ESPERADAS = ["E/S", "CONTA", "SUBCONTA", "FORNECEDOR", "Valor Pago", "Pgto.", "Banco", "OBS"]
NOMES_COLUNAS_POSICIONAIS = COLUNAS_ESPERADAS  # mesma ordem, usado no fallback posicional
LINHAS_BUSCA_CABECALHO = 10


def _nome_valido(nome_arquivo: str, ano: str) -> bool:
    if nome_arquivo.startswith("~$"):
        return False
    if not nome_arquivo.lower().endswith(".xlsx"):
        return False
    nome_norm = re.sub(r"\s+", " ", nome_arquivo.upper())
    return "FLUXO DE CAIXA" in nome_norm and ano in nome_norm and "LIVE" in nome_norm


def descobrir_arquivos(pasta_origem: Path, ano: str) -> list[Path]:
    if not pasta_origem.exists():
        logger.error("Pasta de origem não existe: %s", pasta_origem)
        return []
    encontrados = [
        p for p in pasta_origem.iterdir() if p.is_file() and _nome_valido(p.name, ano)
    ]
    return sorted(encontrados)


def _detectar_inicio_dados(caminho: Path, aba: str) -> tuple[int, list[str] | None, bool]:
    """Retorna (linha_inicial_dados_1indexed, nomes_colunas_ou_None, usou_fallback).

    Procura a linha com os nomes reais de coluna nas primeiras linhas da aba.
    Se não encontrar (planilha sem cabeçalho nomeado), cai para o modo
    posicional: localiza a linha de marcadores ("Coluna2" na 2ª célula) e
    assume que os dados começam logo depois, com colunas fixas por posição.
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        if aba not in wb.sheetnames:
            return -1, None, False
        ws = wb[aba]
        linha_marcador = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=LINHAS_BUSCA_CABECALHO, values_only=True), start=1):
            valores = [str(v).strip() if v is not None else "" for v in row[: len(COLUNAS_ESPERADAS)]]
            if valores == COLUNAS_ESPERADAS:
                return i + 1, None, False
            if len(row) > 1 and row[1] == "Coluna2":
                linha_marcador = i

        if linha_marcador is not None:
            return linha_marcador + 1, NOMES_COLUNAS_POSICIONAIS, True

        return -1, None, False
    finally:
        wb.close()


def _ler_planilha(caminho: Path, cfg: Config) -> pd.DataFrame | None:
    linha_inicio, nomes_colunas, usou_fallback = _detectar_inicio_dados(caminho, cfg.aba)
    if linha_inicio == -1:
        logger.warning(
            "Arquivo ignorado (cabeçalho da aba '%s' não reconhecido): %s", cfg.aba, caminho.name
        )
        return None

    if usou_fallback:
        logger.warning(
            "Arquivo '%s': cabeçalho nomeado não encontrado, usando posição fixa de colunas "
            "(dados a partir da linha %d).",
            caminho.name,
            linha_inicio,
        )

    try:
        if nomes_colunas:
            df = pd.read_excel(
                caminho,
                sheet_name=cfg.aba,
                skiprows=linha_inicio - 1,
                header=None,
                names=nomes_colunas,
                usecols="A:H",
                engine="openpyxl",
            )
        else:
            # linha_inicio é a 1ª linha de DADOS; o cabeçalho nomeado está uma
            # linha antes dela, então pulamos até a linha do cabeçalho
            # (exclusive) e deixamos o pandas consumi-la como header=0.
            df = pd.read_excel(
                caminho,
                sheet_name=cfg.aba,
                skiprows=linha_inicio - 2,
                usecols="A:H",
                engine="openpyxl",
            )
    except Exception:
        logger.exception("Falha ao ler arquivo %s", caminho.name)
        return None

    faltantes = set(COLUNAS_ESPERADAS) - set(df.columns)
    if faltantes:
        logger.warning(
            "Arquivo ignorado (colunas ausentes %s): %s", sorted(faltantes), caminho.name
        )
        return None

    return df


def _aplicar_filtros(df: pd.DataFrame, cfg: Config, nome_arquivo: str) -> pd.DataFrame:
    total_lido = len(df)

    df = df.dropna(how="all")

    conta = df["CONTA"].astype(str).str.strip()
    subconta = df["SUBCONTA"].astype(str).str.strip()
    banco_bruto = df["Banco"].astype(str).str.strip()

    mascara_rn01 = conta.str.startswith(cfg.conta_prefixo) & subconta.str.startswith(
        tuple(cfg.subconta_prefixos)
    )
    apos_rn01 = int(mascara_rn01.sum())

    banco_norm = banco_bruto.str.upper().str.replace(r"\s+", "", regex=True)
    banco_aceito_norm = cfg.banco_aceito.upper().replace(" ", "")

    bancos_no_rn01 = banco_bruto[mascara_rn01].value_counts()
    if not bancos_no_rn01.empty:
        logger.info(
            "Arquivo '%s': bancos encontrados dentro do filtro RN-01: %s",
            nome_arquivo,
            dict(bancos_no_rn01),
        )

    mascara_rn08 = mascara_rn01 & (banco_norm == banco_aceito_norm)
    apos_rn08 = int(mascara_rn08.sum())

    df_filtrado = df.loc[mascara_rn08].copy()

    datas = pd.to_datetime(df_filtrado["Pgto."], dayfirst=True, errors="coerce")
    linhas_sem_data = int(datas.isna().sum())
    df_filtrado = df_filtrado.loc[datas.notna()].copy()
    df_filtrado["_data_pagamento"] = datas.loc[datas.notna()].dt.strftime("%Y-%m-%d")

    df_filtrado["_valor_centavos"] = (
        df_filtrado["Valor Pago"].astype(float).abs().round(2) * 100
    ).round().astype("int64")

    logger.info(
        "Arquivo '%s': %d linhas lidas | %d após RN-01 (conta/subconta) | "
        "%d após RN-08 (banco) | %d descartadas por data inválida | %d linhas finais",
        nome_arquivo,
        total_lido,
        apos_rn01,
        apos_rn08,
        linhas_sem_data,
        len(df_filtrado),
    )

    return df_filtrado


def _linhas_para_gravar(df: pd.DataFrame, nome_arquivo: str) -> list[dict[str, Any]]:
    registros = []
    for _, row in df.iterrows():
        cliente = str(row["FORNECEDOR"]).strip()
        subconta = str(row["SUBCONTA"]).strip()
        banco = str(row["Banco"]).strip()
        obs = row.get("OBS")
        obs = None if pd.isna(obs) else str(obs).strip()

        registros.append(
            {
                "cliente": cliente,
                "cliente_normalizado": db.normalizar_texto(cliente),
                "data_pagamento": row["_data_pagamento"],
                "valor_centavos": int(row["_valor_centavos"]),
                "banco": banco,
                "subconta": subconta,
                "obs": obs,
                "arquivo_origem": nome_arquivo,
            }
        )
    return registros


def executar(cfg: Config | None = None) -> dict[str, Any]:
    inicio = time.time()
    cfg = cfg or carregar_config()

    arquivos = descobrir_arquivos(cfg.pasta_origem, cfg.ano)
    logger.info("Pasta de origem: %s", cfg.pasta_origem)
    logger.info("%d arquivo(s) encontrado(s) para o ano %s.", len(arquivos), cfg.ano)

    todas_as_linhas: list[dict[str, Any]] = []
    processados = 0
    ignorados = 0

    for caminho in arquivos:
        df = _ler_planilha(caminho, cfg)
        if df is None:
            ignorados += 1
            continue
        df_filtrado = _aplicar_filtros(df, cfg, caminho.name)
        todas_as_linhas.extend(_linhas_para_gravar(df_filtrado, caminho.name))
        processados += 1

    duracao = time.time() - inicio

    stats = {
        "arquivos_processados": processados,
        "arquivos_ignorados": ignorados,
        "linhas_importadas": len(todas_as_linhas),
        "duracao_segundos": round(duracao, 2),
    }

    total_gravado = db.gravar_carga_atomica(cfg.caminho_db, todas_as_linhas, stats)

    logger.info(
        "Sincronização concluída: %d arquivos processados, %d ignorados, "
        "%d linhas gravadas em %.2fs.",
        processados,
        ignorados,
        total_gravado,
        duracao,
    )

    return {**stats, "linhas_gravadas": total_gravado}
